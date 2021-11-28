import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import openpyxl 
import re

# path to parent folder containing the folders with our data
#################### input
directory = r'C:\Users\jrbrz\Desktop\690\assignments'  ###### INPUT PATH that holds folder with historical files

# the name of the folder the raw data files are in
###################### input
faaFolder = '\\FAAdataOriginal'  #####  INPUT PATH for the folder inside the first path
folder = faaFolder[1:] + "\\"

# blank "master" df to append all files into for analysis
faa = pd.DataFrame(columns=['Date', 'Event', 'City', 'State'])

### loop through the FAA data to merge it to a master d
# looping through each raw file, cleaning up encoding issues
# and appending it to the master df
for entry in os.scandir(directory + faaFolder):
    try:
        # creating the file path
        file = "FAAdataOriginal\\" + entry.name

        # loading the file into the openpyxl handler
        wb = load_workbook(file)
        wb.iso_dates = True
        # grabbing actual worksheet name in the xlsx file
        for s in range(len(wb.sheetnames)):
            sheet = wb.sheetnames[s]

        # activating the worksheet
        ws = wb[sheet]

        # grabbing the data from the worksheet
        data = ws.values

        # grabbing the column header names from the file
        columns = next(data)[0:]
        col = []
        for i in columns:
            col.append(str(i).replace("'", "").strip())

        # moving the data into a data frame
        df = pd.DataFrame(data, columns=col)

        df.rename(columns={"Day of Sighting": "Date", "Summary": "Event",
                           "City": "City", "State": "State"}, inplace=True)  # "'Date":"Date",
        # "'City":"City","'State":"State","'Summary":"Event"}

        df = df[['Date', 'Event', 'City', 'State']]
        df['City'] = df['City'].replace("'", "")
        df['State'] = df['State'].replace("'", "")
        df['Date'] = df['Date'].replace("'", "")
        ### cleaning the Event text for encoding oddities
        # strips most special characters
        pat = r'[^a-zA-z0-9.,!?/:;\"\'\s]'
        df['Event'] = df['Event'].apply(lambda x: re.sub(pat, '', x))
        # strips ASCII carriage return
        df['Event'] = df['Event'].apply(lambda x: x.replace("_x000D_", ""))
        # strips the newline characters
        df['Event'] = df['Event'].apply(lambda x: x.replace("\n", " "))
        # strips newline or spaces from the start of a string.
        df['Event'] = df['Event'].apply(lambda x: x.lstrip())
        # strips special character
        df['Event'] = df['Event'].apply(lambda x: x.replace("\xa0", " "))

        # a few text cleaning items
        df['Event'] = df['Event'].str.upper()
        df['City'] = df['City'].str.strip().str.upper()
        df['State'] = df['State'].str.strip().str.upper()

        # final filters to clean data worthless entries
        df = df[~df['Event'].str.contains("DUPLICATE")]
        df = df[df['Date'] != "#VALUE!"]

        # add the file name for use in debugging
        ###df['File'] = entry.name

        # added the df from each individual file into the master df
        faa = faa.append(df, ignore_index=True).fillna(np.NaN)

    # prints the file name that failed to clean
    except:
        print(entry.name)
faa.to_csv("DATA_FAA_merged.csv")